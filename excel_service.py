"""
Servicio de exportación de métricas a Excel
"""
import pandas as pd
from typing import List, Dict
from datetime import datetime
from config import OUTPUT_FILE


class ExcelExportService:
    """Servicio para exportar métricas a archivos Excel"""
    
    def __init__(self, output_file: str = OUTPUT_FILE):
        self.output_file = output_file
    
    def export_metrics(self, metrics: List[Dict]) -> bool:
        """
        Exporta métricas a un archivo Excel con formato profesional
        
        Args:
            metrics: Lista de diccionarios con métricas
            
        Returns:
            True si la exportación fue exitosa
        """
        if not metrics:
            print("⚠️ No hay métricas para exportar")
            return False
        
        try:
            # Crear DataFrame
            df = pd.DataFrame(metrics)
            
            # Ordenar columnas de forma lógica
            priority_columns = [
                'plataforma', 'tipo', 'fecha_creacion', 'titulo', 'mensaje',
                'descripcion', 'url', 'vistas', 'likes', 'comentarios_count',
                'compartidos', 'duracion', 'canal', 'id'
            ]
            
            # Organizar columnas: primero las prioritarias, luego el resto
            available_priority = [col for col in priority_columns if col in df.columns]
            other_columns = [col for col in df.columns if col not in priority_columns]
            df = df[available_priority + other_columns]
            
            # Crear writer de Excel con formato
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Métricas', index=False)
                
                # Obtener el worksheet para aplicar formato
                worksheet = writer.sheets['Métricas']
                
                # Ajustar ancho de columnas
                for idx, col in enumerate(df.columns, 1):
                    # Calcular ancho basado en el contenido
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    # Limitar ancho máximo
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[chr(64 + idx)].width = adjusted_width
                
                # Aplicar formato a encabezados
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_alignment = Alignment(horizontal="center", vertical="center")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # Congelar primera fila
                worksheet.freeze_panes = "A2"
            
            print(f"\n{'='*60}")
            print(f"✅ EXPORTACIÓN EXITOSA")
            print(f"{'='*60}")
            print(f"📁 Archivo: {self.output_file}")
            print(f"📊 Total de registros: {len(metrics)}")
            print(f"📅 Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"{'='*60}\n")
            
            return True
            
        except Exception as e:
            print(f"❌ Error al exportar a Excel: {e}")
            return False
    
    def export_separate_sheets(self, facebook_metrics: List[Dict], youtube_metrics: List[Dict]) -> bool:
        """
        Exporta métricas en hojas separadas por plataforma
        
        Args:
            facebook_metrics: Métricas de Facebook
            youtube_metrics: Métricas de YouTube
            
        Returns:
            True si la exportación fue exitosa
        """
        try:
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                
                # Hoja de Facebook
                if facebook_metrics:
                    df_fb = pd.DataFrame(facebook_metrics)
                    df_fb.to_excel(writer, sheet_name='Facebook', index=False)
                    self._format_worksheet(writer.sheets['Facebook'], df_fb)
                
                # Hoja de YouTube
                if youtube_metrics:
                    df_yt = pd.DataFrame(youtube_metrics)
                    df_yt.to_excel(writer, sheet_name='YouTube', index=False)
                    self._format_worksheet(writer.sheets['YouTube'], df_yt)
                
                # Hoja consolidada
                all_metrics = facebook_metrics + youtube_metrics
                if all_metrics:
                    df_all = pd.DataFrame(all_metrics)
                    df_all.to_excel(writer, sheet_name='Consolidado', index=False)
                    self._format_worksheet(writer.sheets['Consolidado'], df_all)
            
            print(f"\n{'='*60}")
            print(f"✅ EXPORTACIÓN EXITOSA (HOJAS SEPARADAS)")
            print(f"{'='*60}")
            print(f"📁 Archivo: {self.output_file}")
            print(f"📊 Facebook: {len(facebook_metrics)} registros")
            print(f"📊 YouTube: {len(youtube_metrics)} registros")
            print(f"📊 Total: {len(facebook_metrics) + len(youtube_metrics)} registros")
            print(f"📅 Fecha de generación: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"{'='*60}\n")
            
            return True
            
        except Exception as e:
            print(f"❌ Error al exportar a Excel: {e}")
            return False
    
    def _format_worksheet(self, worksheet, df: pd.DataFrame):
        """
        Aplica formato a una hoja de Excel
        
        Args:
            worksheet: Hoja de Excel de openpyxl
            df: DataFrame con los datos
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Ajustar ancho de columnas
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(col)
            )
            adjusted_width = min(max_length + 2, 50)
            column_letter = worksheet.cell(row=1, column=idx).column_letter
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Formato de encabezados
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Congelar primera fila
        worksheet.freeze_panes = "A2"